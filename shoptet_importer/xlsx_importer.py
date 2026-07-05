from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .ai_content import generate_basic_content
from .csv_importer import COLUMN_ALIASES, normalize_header, parse_float
from .product_model import UniversalProduct


def map_headers(headers: list[str]) -> dict[str, int]:
    normalized = {normalize_header(str(h)): i for i, h in enumerate(headers)}
    result: dict[str, int] = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                result[field] = normalized[alias]
                break
    return result


def cell(row: list[Any], index: int | None, default: Any = "") -> Any:
    if index is None:
        return default
    if index >= len(row):
        return default
    return row[index]


def import_xlsx_products(path: str | Path, default_supplier: str = "", default_manufacturer: str = "", limit: int | None = None) -> tuple[list[UniversalProduct], list[dict]]:
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        headers = [str(x or "").strip() for x in next(rows_iter)]
    except StopIteration:
        return [], [{"reason": "empty_file", "file": str(path)}]

    header_map = map_headers(headers)
    products: list[UniversalProduct] = []
    skipped: list[dict] = []

    for row_number, values in enumerate(rows_iter, start=2):
        row = list(values)
        code = str(cell(row, header_map.get("code"), "") or "").strip()
        name = str(cell(row, header_map.get("name"), "") or "").strip()
        if not code or not name:
            skipped.append({"row": row_number, "reason": "missing_code_or_name"})
            continue

        purchase = parse_float(cell(row, header_map.get("purchase_price"), 0))
        sale = parse_float(cell(row, header_map.get("sale_price"), 0))
        if sale <= 0:
            sale = round(purchase * 1.35, 2) if purchase > 0 else 0.0

        product = UniversalProduct(
            code=code,
            name=name,
            supplier=str(cell(row, header_map.get("supplier"), default_supplier) or default_supplier).strip(),
            manufacturer=str(cell(row, header_map.get("manufacturer"), default_manufacturer) or default_manufacturer).strip(),
            supplier_code=code,
            model=str(cell(row, header_map.get("model"), "") or "").strip(),
            ean=str(cell(row, header_map.get("ean"), "") or "").strip(),
            category=str(cell(row, header_map.get("category"), "Produkty") or "Produkty").strip(),
            purchase_price=purchase,
            sale_price=sale,
            standard_price=sale,
            source=f"xlsx:{path.name}",
        )
        products.append(generate_basic_content(product))
        if limit and len(products) >= limit:
            break

    return products, skipped
